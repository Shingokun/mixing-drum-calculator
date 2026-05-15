# ============================================================
# excel_exporter.py — Xuất báo cáo ra Excel
# ============================================================
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from app.core.session import ProjectSession
from app.ui.i18n import _


# ── Màu sắc chủ đạo ──────────────────────────────────────
C_HEADER_BG  = "1E3A5F"   # Xanh đậm
C_HEADER_FG  = "FFFFFF"   # Trắng
C_SUB_BG     = "2C6FAC"   # Xanh vừa
C_SUB_FG     = "FFFFFF"
C_ROW_ALT    = "EBF3FB"   # Xanh nhạt xen kẽ
C_OK         = "2ECC71"   # Xanh lá
C_FAIL       = "E74C3C"   # Đỏ
C_LABEL_BG   = "D6E4F0"   # Xám xanh nhạt

# ── Tiện ích ─────────────────────────────────────────────
def _side(style="thin"):
    return Side(style=style, color="AAAAAA")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font(bold=True, size=11):
    return Font(name="Calibri", bold=bold, size=size, color=C_HEADER_FG)

def _body_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")


def _write_section_header(ws, row: int, title: str, num_cols: int = 5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _hdr_font(bold=True, size=12)
    cell.fill = _fill(C_HEADER_BG)
    cell.alignment = _center()
    cell.border = _border()


def _write_sub_header(ws, row: int, cols: list):
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _hdr_font(bold=True, size=10)
        cell.fill = _fill(C_SUB_BG)
        cell.alignment = _center()
        cell.border = _border()


def _write_row(ws, row: int, values: list, alt=False):
    bg = C_ROW_ALT if alt else "FFFFFF"
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = _body_font(bold=False, size=10)
        cell.fill = _fill(bg)
        cell.alignment = _right() if isinstance(val, (int, float)) else _left()
        cell.border = _border()


def export_excel(session: ProjectSession, filepath: str):
    wb = Workbook()
    _sheet_summary(wb, session)
    _sheet_uc03(wb, session)
    _sheet_uc04(wb, session)
    _sheet_uc05(wb, session)
    wb.save(filepath)


# ─────────────────────────────────────────────────────────
def _sheet_summary(wb: Workbook, s: ProjectSession):
    ws = wb.active
    ws.title = _("nav_uc06")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.row_dimensions[1].height = 40

    # Tiêu đề lớn
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = _("rpt_title").replace("\n", " ")
    t.font = Font(name="Calibri", bold=True, size=16, color=C_HEADER_FG)
    t.fill = _fill(C_HEADER_BG)
    t.alignment = _center()
    t.border = _border()

    row = 3
    _write_section_header(ws, row, _("rpt_sec_input")); row += 1
    pairs = [
        (_("p_power") + " P",       f"{s.inputs.power_kw:.2f} kW"),
        (_("p_rpm") + " n",         f"{s.inputs.rpm_out:.1f} " + _("unit_rpm")),
        (_("p_life") + " L",           f"{s.inputs.service_life_year:.1f} " + _("unit_year")),
        (_("p_eta_kn"),       f"{s.inputs.eta_kn:.3f}"),
        (_("p_eta_ol"),          f"{s.inputs.eta_ol:.3f}"),
        (_("p_eta_brc"),        f"{s.inputs.eta_brc:.3f}"),
        (_("p_eta_brt"),        f"{s.inputs.eta_brt:.3f}"),
        (_("p_eta_x"),        f"{s.inputs.eta_x:.3f}"),
    ]
    for i, (k, v) in enumerate(pairs):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        lbl = ws.cell(row=row, column=1, value=k)
        val = ws.cell(row=row, column=4, value=v)
        lbl.font = _body_font(bold=True); lbl.fill = _fill(C_LABEL_BG if i%2==0 else "FFFFFF")
        val.font = _body_font(); val.fill = _fill(C_ROW_ALT if i%2==0 else "FFFFFF")
        lbl.alignment = _left(); val.alignment = _right()
        lbl.border = _border(); val.border = _border()
        row += 1

    row += 1
    _write_section_header(ws, row, _("rpt_sec_motor")); row += 1
    m = s.motor
    summary = [
        (_("lbl_total_eta"),         f"{m.eta:.4f}"),
        (_("lbl_req_power"),           f"{m.p_ct_kw:.4f} kW"),
        (_("lbl_motor_sel"),                    m.motor_name or "—"),
        (_("col_p"),             f"{m.rated_power_kw:.1f} kW"),
        (_("col_n"),          f"{m.rated_rpm:.0f} " + _("unit_rpm")),
        (_("lbl_actual_utotal"),            f"{m.u_ch_actual:.4f}"),
        (_("lbl_actual_uh"),     f"{m.u_h_actual:.4f}"),
        (_("p_u1"),        f"{s.inputs.u1:.2f}"),
        (_("lbl_u2"),         f"{m.u2:.2f}"),
        (_("lbl_z_belt"),                       f"{s.belt.num_belts}"),
        (_("lbl_v_belt"),                      f"{s.belt.belt_velocity_ms:.3f} " + _("unit_ms")),
        (_("lbl_mte"),      f"{s.gearbox.cone.m_te:.2f} mm"),
        (_("lbl_z1z2") + " z1",               f"{s.gearbox.cone.z1}"),
        (_("lbl_z1z2") + " z2",               f"{s.gearbox.cone.z2}"),
        (_("lbl_re"),           f"{s.gearbox.cone.Re_mm:.2f} mm"),
        (_("lbl_b_width"),            f"{s.gearbox.cone.b_mm:.2f} mm"),
        (_("lbl_sig_h"),              f"{s.gearbox.cone.sig_H:.2f} MPa"),
        (_("lbl_check_f1"),           f"{s.gearbox.cone.sigma_F1_actual:.2f} / {s.gearbox.cone.sig_F1:.2f} MPa"),
        (_("lbl_check_f2"),           f"{s.gearbox.cone.sigma_F2_actual:.2f} / {s.gearbox.cone.sig_F2:.2f} MPa"),
        (_("g_strength_check"),                       _("rpt_pass") if s.gearbox.strength_ok else _("rpt_fail")),
    ]
    for i, (k, v) in enumerate(summary):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        lbl = ws.cell(row=row, column=1, value=k)
        val = ws.cell(row=row, column=4, value=v)
        bg = C_ROW_ALT if i % 2 == 0 else "FFFFFF"
        lbl.fill = val.fill = _fill(bg)
        lbl.font = _body_font(bold=True)
        if _("rpt_pass") in str(v): val.font = Font(name="Calibri", bold=True, size=10, color=C_OK)
        elif _("rpt_fail") in str(v): val.font = Font(name="Calibri", bold=True, size=10, color=C_FAIL)
        else: val.font = _body_font()
        lbl.alignment = _left(); val.alignment = _right()
        lbl.border = val.border = _border()
        row += 1


def _sheet_uc03(wb: Workbook, s: ProjectSession):
    ws = wb.create_sheet(_("nav_uc03"))
    for col, w in zip("ABCDE", [35, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    row = 1
    _write_section_header(ws, row, _("rpt_shaft_table")); row += 1
    _write_sub_header(ws, row, [_("shaft_param"), _("shaft_motor"), _("shaft_1"), _("shaft_2"), _("shaft_3")]); row += 1

    m = s.motor
    shaft_rows = [
        (_("shaft_power"),
         m.shaft_powers.get('dc','—'), m.shaft_powers.get('I','—'),
         m.shaft_powers.get('II','—'), m.shaft_powers.get('III','—')),
        (_("shaft_speed"),
         m.shaft_rpms.get('dc','—'), m.shaft_rpms.get('I','—'),
         m.shaft_rpms.get('II','—'), m.shaft_rpms.get('III','—')),
        (_("shaft_torque"),
         m.shaft_torques.get('dc','—'), m.shaft_torques.get('I','—'),
         m.shaft_torques.get('II','—'), m.shaft_torques.get('III','—')),
    ]
    for i, row_data in enumerate(shaft_rows):
        _write_row(ws, row, list(row_data), alt=(i%2==0)); row += 1


def _sheet_uc04(wb: Workbook, s: ProjectSession):
    ws = wb.create_sheet(_("nav_uc04"))
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25

    row = 1
    _write_section_header(ws, row, _("rpt_sec_belt"), num_cols=2); row += 1
    b = s.belt
    params = [
        (_("lbl_d1"),     f"{b.d1_mm:.1f} mm"),
        (_("lbl_d2"),      f"{b.d2_mm:.1f} mm"),
        (_("lbl_v_belt"),                   f"{b.belt_velocity_ms:.3f} " + _("unit_ms")),
        (_("lbl_l_belt"),                 f"{b.belt_length_mm:.0f} mm"),
        (_("lbl_a_belt"),              f"{b.center_distance_mm:.1f} mm"),
        (_("lbl_alpha1"),        f"{b.alpha1_deg:.2f}°"),
        (_("lbl_z_belt"),                    str(b.num_belts)),
        (_("lbl_ft"),                    f"{b.Ft_N:.1f} N"),
        (_("lbl_f0"),            f"{b.F0_N:.1f} N"),
        (_("rpt_result"),             _("rpt_pass") if b.velocity_ok else _("rpt_fail")),
    ]
    for i, (k, v) in enumerate(params):
        lbl = ws.cell(row=row, column=1, value=k)
        val = ws.cell(row=row, column=2, value=v)
        bg = C_ROW_ALT if i%2==0 else "FFFFFF"
        lbl.fill = val.fill = _fill(bg)
        lbl.font = _body_font(bold=True)
        if _("rpt_pass") in str(v): val.font = Font(name="Calibri", bold=True, size=10, color=C_OK)
        elif _("rpt_fail") in str(v): val.font = Font(name="Calibri", bold=True, size=10, color=C_FAIL)
        else: val.font = _body_font()
        lbl.alignment = _left(); val.alignment = _right()
        lbl.border = val.border = _border()
        row += 1


def _sheet_uc05(wb: Workbook, s: ProjectSession):
    ws = wb.create_sheet(_("nav_uc05"))
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25

    row = 1
    _write_section_header(ws, row, _("rpt_sec_gear"), num_cols=2); row += 1
    c = s.gearbox.cone
    params = [
        (_("lbl_u_gear"),       f"{c.u_tt:.4f}"),
        (_("lbl_delta_u"),         f"{c.delta_u_pct:.2f}%"),
        (_("lbl_mte"),     f"{c.m_te:.2f} mm"),
        (_("lbl_z1z2") + " z₁",             str(c.z1)),
        (_("lbl_z1z2") + " z₂",             str(c.z2)),
        (_("lbl_re"),          f"{c.Re_mm:.2f} mm"),
        ("d_e1", f"{c.de1_mm:.2f} mm"),
        (_("lbl_b_width"),           f"{c.b_mm:.2f} mm"),
        ("δ₁",             f"{c.delta1_deg:.3f}°"),
        ("δ₂",             f"{c.delta2_deg:.3f}°"),
        (_("lbl_sig_h"),          f"{c.sig_H:.2f} MPa"),
        (_("lbl_sig_f1") + " / actual",          f"{c.sig_F1:.2f} / {c.sigma_F1_actual:.2f} MPa"),
        (_("lbl_sig_f2") + " / actual",          f"{c.sig_F2:.2f} / {c.sigma_F2_actual:.2f} MPa"),
        (_("lbl_ft"),                     f"{c.Ft_N:.1f} N"),
        (_("lbl_fr"),               f"{c.Fr_N:.1f} N"),
        (_("lbl_fa"),                f"{c.Fa_N:.1f} N"),
        (_("g_strength_check"),                     _("rpt_pass") if s.gearbox.strength_ok else _("rpt_fail")),
    ]
    for i, (k, v) in enumerate(params):
        lbl = ws.cell(row=row, column=1, value=k)
        val = ws.cell(row=row, column=2, value=v)
        bg = C_ROW_ALT if i%2==0 else "FFFFFF"
        lbl.fill = val.fill = _fill(bg)
        lbl.font = _body_font(bold=True)
        if _("rpt_pass") in str(v): val.font = Font(name="Calibri", bold=True, size=10, color=C_OK)
        elif _("rpt_fail") in str(v): val.font = Font(name="Calibri", bold=True, size=10, color=C_FAIL)
        else: val.font = _body_font()
        lbl.alignment = _left(); val.alignment = _right()
        lbl.border = val.border = _border()
        row += 1
